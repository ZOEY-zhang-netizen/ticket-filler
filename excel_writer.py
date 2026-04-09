"""Write extracted ticket products into template.xlsx via openpyxl."""

from __future__ import annotations

from copy import copy
from datetime import datetime
import re

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

CHANNELS = [
    ('抖团-门票', '期票', '期票', '提前9小时', 1, 'tuan'),
    ('抖音半直连', '日历票', '日历票', '提前1天', 5, 'semi'),
]
AH_AJ_LIST = '游玩当日,每1天,每5天,每30天,每60天,每90天,每365天'
RED_FILL = PatternFill(fill_type='solid', fgColor='FFFF0000')
DATE_FMT = 'm"月"d"日"'


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.strptime(str(value), '%Y-%m-%d')


def _month_day(value) -> str:
    if not value:
        return '（请填写）'
    if isinstance(value, str):
        value = _parse_date(value)
    return f'{value.month}月{value.day}日'


def _day_span(start: str | None, end: str | None) -> int | None:
    s = _parse_date(start)
    e = _parse_date(end)
    if not s or not e:
        return None
    return max((e - s).days, 0)


def ladder_value(start: str | None, end: str | None) -> str:
    span = _day_span(start, end)
    if span is None:
        return '每30天'
    if span == 0:
        return '游玩当日'
    if span == 1:
        return '每1天'
    if 2 <= span <= 5:
        return '每5天'
    if 6 <= span <= 30:
        return '每30天'
    if 31 <= span <= 60:
        return '每60天'
    if 61 <= span <= 90:
        return '每90天'
    return '每365天'


def _copy_row_style(ws, src_row: int, dst_row: int):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = copy(src.number_format)
        if src.font:
            dst.font = copy(src.font)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if isinstance(src.value, str) and src.value.startswith('='):
            try:
                dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = src.value
        elif col in {9, 20, 21, 22, 23}:
            dst.value = src.value


def _clear_target_rows(ws):
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)


def _extend_existing_validations(ws, end_row: int = 2000):
    def extend_sqref(sqref: str) -> str:
        parts = []
        for part in str(sqref).split():
            m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', part)
            if m:
                parts.append(f'{m.group(1)}2:{m.group(3)}{end_row}')
            else:
                parts.append(part)
        return ' '.join(parts)
    for dv in ws.data_validations.dataValidation:
        dv.sqref = extend_sqref(dv.sqref)
    dv_period = DataValidation(type='list', formula1=f'"{AH_AJ_LIST}"', allow_blank=True)
    ws.add_data_validation(dv_period)
    dv_period.add(f'AH2:AH{end_row}')
    dv_period.add(f'AJ2:AJ{end_row}')


def _set_date(ws, cell_ref: str, value):
    cell = ws[cell_ref]
    if value:
        cell.value = value
        cell.number_format = DATE_FMT
    else:
        cell.value = None


def _mark_review(ws, cell_ref: str):
    ws[cell_ref].fill = copy(RED_FILL)


def _wrap_row(ws, row: int):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        old = cell.alignment or Alignment()
        cell.alignment = Alignment(
            wrap_text=True,
            horizontal=old.horizontal,
            vertical=old.vertical,
            text_rotation=old.text_rotation,
            indent=old.indent,
            shrink_to_fit=old.shrink_to_fit,
        )


def _ensure_sentence(text: str | None) -> str:
    if not text:
        return '（请填写）'
    return str(text).strip().rstrip('；;。')


def _strip_rule_prefix(text: str | None) -> str:
    text = _ensure_sentence(text)
    return re.sub(r'^(?:POI|抖音平台及直播间|抖音平台及美团直播间|抖音直播间|抖音、美团|其他渠道|直播间)[：:]', '', text).strip()


def _fmt_cn_date(value) -> str:
    if not value:
        return '（请填写）'
    if isinstance(value, str):
        value = _parse_date(value)
    return f'{value.year}年{value.month}月{value.day}日'


def _build_use_rule(product: dict, channel_key: str) -> str:
    entry = product.get('entry') or '刷身份证入园'
    entry = entry.replace('刷身份证', '刷本人购票身份证')

    if channel_key == 'semi':
        return f'{entry}；'

    start = product.get('use_start')
    end = product.get('use_end')
    if start and end:
        period = f'{_fmt_cn_date(start)}-{_fmt_cn_date(end)}期间仅限入园使用1次；'
    elif product.get('use_rule_raw'):
        raw = _ensure_sentence(product.get('use_rule_raw'))
        period = f'{raw}；'
    else:
        period = '使用时间以页面展示为准；'

    return f'本产品{period}{entry}（门票仅限入园当日有效，离开园区后当日再入园需重新购票）；'


def _build_desc(product: dict, channel_key: str) -> str:
    purchase_rule = product.get('purchase_rule_tuan') if channel_key == 'tuan' else product.get('purchase_rule_semi')
    purchase_rule = _strip_rule_prefix(purchase_rule)
    crowd = _ensure_sentence(product.get('crowd'))
    refund = _ensure_sentence(product.get('refund'))
    return "\n".join([
        f'【费用包含】：上海海昌海洋公园{product["fee_desc"]}；',
        f'【游玩人群】：{crowd}；',
        f'【购买规则】：{purchase_rule}；',
        f'【使用规则】：{_build_use_rule(product, channel_key)}',
        f'【退改规则】：{refund}；',
    ])


def _write_row(ws, row: int, product: dict, channel: tuple):
    sale_start = _parse_date(product['sale_start'])
    sale_end = _parse_date(product['sale_end'])
    use_start = _parse_date(product['use_start'])
    use_end = _parse_date(product.get('use_end')) if product.get('use_end') else None

    channel_name, platform, ticket_type, advance, order_limit, channel_key = channel
    id_period = '游玩当日' if channel_key == 'semi' else ladder_value(product['sale_start'], product['sale_end'])
    id_limit = product['id_limit_tuan'] if channel_key == 'tuan' else product['id_limit_semi']
    phone_limit = product.get('phone_limit_tuan') if channel_key == 'tuan' else product.get('phone_limit_semi')
    phone_period = id_period if phone_limit is not None else None

    ws[f'C{row}'] = channel_name
    ws[f'D{row}'] = product.get('year', datetime.now().year)
    ws[f'E{row}'] = '上海'
    ws[f'F{row}'] = None
    _mark_review(ws, f'F{row}')

    ws[f'G{row}'] = product['name']
    ws[f'H{row}'] = '（通用）'
    _set_date(ws, f'J{row}', sale_start)
    _set_date(ws, f'K{row}', sale_end)
    _set_date(ws, f'L{row}', use_start)
    if product.get('use_manual_review'):
        _set_date(ws, f'M{row}', use_end)
        _mark_review(ws, f'L{row}')
        _mark_review(ws, f'M{row}')
    elif product.get('is_afternoon'):
        _set_date(ws, f'M{row}', use_end)
        if product.get('use_rule_raw') and any(k in product['use_rule_raw'] for k in ('根据', '开放时间', '①', '②', '③', '④')):
            _mark_review(ws, f'L{row}')
            _mark_review(ws, f'M{row}')
    else:
        _set_date(ws, f'M{row}', use_end)

    ws[f'N{row}'] = product['list_price']
    ws[f'O{row}'] = product['sale_price']
    ws[f'P{row}'] = '否'
    ws[f'Q{row}'] = '不限'
    ws[f'R{row}'] = '不限'
    ws[f'S{row}'] = platform
    ws[f'X{row}'] = '-'
    ws[f'AB{row}'] = '是'
    ws[f'AC{row}'] = ticket_type
    ws[f'AD{row}'] = advance
    ws[f'AE{row}'] = product.get('age_limit') or None
    ws[f'AG{row}'] = order_limit
    ws[f'AH{row}'] = id_period
    ws[f'AI{row}'] = id_limit
    ws[f'AJ{row}'] = phone_period
    ws[f'AK{row}'] = phone_limit
    ws[f'AL{row}'] = '否'
    ws[f'AM{row}'] = _build_desc(product, channel_key)
    ws[f'AN{row}'] = product['refund']
    _wrap_row(ws, row)


def write_excel(products: list[dict], template_path: str, output_path: str) -> str:
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    _clear_target_rows(ws)
    row = 2
    for product in products:
        for idx, channel in enumerate(CHANNELS):
            template_row = 2 if idx == 0 else 3
            if row > ws.max_row:
                ws.insert_rows(row)
            _copy_row_style(ws, template_row, row)
            _write_row(ws, row, product, channel)
            row += 1
    _extend_existing_validations(ws, 2000)
    ws.freeze_panes = 'A2'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)
    return output_path
